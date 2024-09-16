from tkinter import *
from datetime import date
from tkinter import filedialog, messagebox
from tkinter.ttk import Combobox
from PIL import Image, ImageTk
import os
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as oImage
import pathlib
from tkcalendar import Calendar
import re
from docx import Document
from docx.shared import Inches
from docx.oxml import OxmlElement
from docx.oxml.ns import qn
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from googleapiclient.discovery import build
from google.oauth2 import service_account
from googleapiclient.http import MediaFileUpload
import win32com.client
from PIL import Image, ImageTk


background = "#06283D"
framebg = "#F0F0F0"
framefg = "#06283D"

root = Tk()
root.title("CEC Student Registration Form")
root.geometry("1250x700+210+100")
root.config(bg=background)

file = pathlib.Path('Student_data1.xlsx')
if not file.exists():
    file = Workbook()
    sheet = file.active
    sheet['A1'] = "Registration No"
    sheet['B1'] = "Name"
    sheet['C1'] = "Course Title"
    sheet['D1'] = "Gender"
    sheet['E1'] = "Students DOB"
    sheet['F1'] = "Date"
    sheet['G1'] = "Email Id"
    sheet['H1'] = "Students Mobile No"
    sheet['I1'] = "Parent Contact No"
    sheet['J1'] = "Qualification"
    sheet['K1'] = "Address"
    file.save('Student_data1.xlsx')




def authenticate():
    SCOPES = ['https://www.googleapis.com/auth/drive']
    SERVICE_ACCOUNT_FILE = 'C:/Users/abin/Downloads/adroit-crow-427515-q4-3dc7f14c6bd2.json'
    creds = service_account.Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)
    return creds

def upload_pdf(file_path, student_first_name):
    creds = authenticate()
    service = build('drive', 'v3', credentials=creds)

    file_metadata = {
        'name': f'{student_first_name}.pdf',
        'parents': ["1znbCiV3HL_jEP5rplm3_UXWc6FJxuZEt"]
    }

    media = MediaFileUpload(file_path, mimetype='application/pdf')

    file = service.files().create(
        body=file_metadata,
        media_body=media,
        fields='id'
    ).execute()

    print(f'File ID: {file.get("id")}')
    messagebox.showinfo("Success", f"The PDF has been saved successfully on Google Drive.")



def Exit():
    root.destroy()

def Clear():
    global img
    Name.set('')
    Quali_student.set('')
    Address_text.delete('1.0', END)
    DOB.set('')
    Email_id.set('')
    Student_Mobile.set('')
    Parent_Contact.set('')
    registration_no()
    course_combobox.set("Select a course")
    gender.set("Male")
    saveButton.config(state='normal')

    img = Image.open("C:/Users/abin/OneDrive/Desktop/F/REimages/upload photo.png")
    resized_image = img.resize((190, 190))
    photo = ImageTk.PhotoImage(resized_image)
    frame1_lbl.config(image=photo)
    frame1_lbl.image = photo

    img = Image.open("C:/Users/abin/OneDrive/Desktop/F/REimages/signature_logo_round 1.png")
    resized_image = img.resize((190, 190))
    photo = ImageTk.PhotoImage(resized_image)
    frame2_lbl.config(image=photo)
    frame2_lbl.image = photo

    img = Image.open("C:/Users/abin/OneDrive/Desktop/F/REimages/Aadhar1.png")
    resized_image = img.resize((190, 190))
    photo = ImageTk.PhotoImage(resized_image)
    frame3_lbl.config(image=photo)
    frame3_lbl.image = photo

    img = Image.open("C:/Users/abin/OneDrive/Desktop/F/REimages/electricity bill.png")
    resized_image = img.resize((590, 835))
    photo = ImageTk.PhotoImage(resized_image)
    frame4_lbl.config(image=photo)
    frame4_lbl.image = photo

def validate_email(email):
    pattern = r'^[a-zA-Z0-9._%+-]+@gmail\.com$'
    return re.match(pattern, email)

def validate_mobile(number):
    return re.match(r'^\d{10}$', number)

def Save():
    global N1
    R1 = Registration.get()
    N1 = Name.get()
    Q1 = Quali_student.get()
    try:
        G1 = gender.get()
    except:
        messagebox.showerror("error", "Select Gender!")
    D2 = DOB.get()
    D1 = Date.get()
    A1 = Address_text.get("1.0", END).strip()
    E1 = Email_id.get()
    S1 = Student_Mobile.get()
    P1 = Parent_Contact.get()
    C1 = course_combobox.get()

    if not N1 or not Q1 or not D2 or not A1 or not E1 or not S1 or not P1 or C1 == "Select a course":
        messagebox.showerror("Error", "All fields are required!")
        return

    if not all(x.isalpha() or x.isspace() for x in N1):
        messagebox.showerror("Error", "Name cannot contain numbers!")
        return

    if not validate_mobile(S1):
        messagebox.showerror("Error", "Student's Mobile No should contain exactly 10 digits and no alphabets!")
        return

    if not validate_mobile(P1):
        messagebox.showerror("Error", "Parent's Contact No should contain exactly 10 digits and no alphabets!")
        return

    if not validate_email(E1):
        messagebox.showerror("Error", "Enter a valid Email ID in the format: example@gmail.com!")
        return
    
    file = openpyxl.load_workbook('Student_data1.xlsx')
    sheet = file.active
    new_row = sheet.max_row + 1
    sheet.cell(column=1, row=new_row, value=R1)
    sheet.cell(column=2, row=new_row, value=N1)
    sheet.cell(column=3, row=new_row, value=C1)
    sheet.cell(column=4, row=new_row, value=G1)
    sheet.cell(column=5, row=new_row, value=D2)
    sheet.cell(column=6, row=new_row, value=D1)
    sheet.cell(column=7, row=new_row, value=E1)
    sheet.cell(column=8, row=new_row, value=S1)
    sheet.cell(column=9, row=new_row, value=P1)
    sheet.cell(column=10, row=new_row, value=Q1)
    sheet.cell(column=11, row=new_row, value=A1)

    file.save(r'Student_data1.xlsx')

    # Save images to the 'images' folder with student's name
    save_images(N1)

    wb = load_workbook("C:/Users/abin/OneDrive/Desktop/F/Registration Form CEC.xlsx")
    ws1 = wb.worksheets[0]

    assign_value = {'D26': R1, 'D27': D1, 'D28': C1, 'D29': N1, 'D30': G1, 'D31': D2, 'D32': S1, 'D33': P1,
                    'D34': E1, 'D35': Q1, 'D36': A1}
        
    for i in assign_value:
        if assign_value[i] != '-':
            ws1[f'{i}'] = assign_value[i]


    #Adding images

    wb['Sheet1']._images
    print(wb['Sheet1']._images)
    del wb['Sheet1']._images[0:]

    img_path = f"C:/Users/abin/OneDrive/Desktop/F/REimages/College name.png"
    cell = 'A1' 
    img = oImage(img_path)
    img.width = 710
    img.height = 90
    ws1.add_image(img,cell)


    img_path = f"C:/Users/abin/OneDrive/Desktop/F/REimages/College name.png"
    cell = 'I1' 
    img = oImage(img_path)
    img.width = 710
    img.height = 90
    ws1.add_image(img,cell)


    img_path = f"C:/Users/abin/OneDrive/Desktop/F/REimages/College name.png"
    cell = 'Q1' 
    img = oImage(img_path)
    img.width = 710
    img.height = 90
    ws1.add_image(img,cell)


    img_path = f"C:/Users/abin/OneDrive/Desktop/F/images/{N1}/{N1}_photo.png"
    cell = 'D13' 
    img = oImage(img_path)
    img.width = 138
    img.height = 139
    ws1.add_image(img,cell)


    img_path = f"C:/Users/abin/OneDrive/Desktop/F/images/{N1}/{N1}_signature.png"
    cell = 'D22' 
    img = oImage(img_path)
    img.width = 131
    img.height = 48
    ws1.add_image(img,cell)


    img_path = f"C:/Users/abin/OneDrive/Desktop/F/images/{N1}/{N1}_aadhar.png"
    cell = 'J19' 
    img = oImage(img_path)
    img.width = 500
    img.height = 300
    ws1.add_image(img,cell)


    img_path = f"C:/Users/abin/OneDrive/Desktop/F/images/{N1}/{N1}_bill.png"
    cell = 'R6' 
    img = oImage(img_path)
    img.width = 595.28
    img.height = 841.89
    ws1.add_image(img,cell)

        
    wb.save("Registration Form CEC.xlsx")
    wb.close()

    # Call the excel_to_pdf function
    Clear()
    registration_no()

    first_name = N1.split()[0].lower()
    pdf_path=excel_to_pdf("C:/Users/abin/OneDrive/Desktop/F/Registration Form CEC.xlsx",f"C:/Users/abin/OneDrive/Desktop/F/REimages/PDFs/{first_name}.pdf")

    # Upload the PDF to Google Drive
    upload_pdf(pdf_path, first_name)

    messagebox.showinfo("Success", "Registration form has been filled successfully.\nThe PDF has been saved successfully in the folder and on Google Drive.")

def add_image_to_excel(ws, img_path, cell, width, height):
    if os.path.exists(img_path):
        img = oImage(img_path)
        img.width = width
        img.height = height
        ws.add_image(img, cell)

def save_images(student_name):
    images_folder = 'C:/Users/abin/OneDrive/Desktop/F/images'
    student_folder = os.path.join(images_folder, student_name)

    if not os.path.exists(student_folder):
        os.makedirs(student_folder)

    image_info = [(frame1_lbl, 'photo'), (frame2_lbl, 'signature'), (frame3_lbl, 'aadhar'), (frame4_lbl, 'bill')]
    for lbl, doc_type in image_info:
        if lbl.image:
            filename = f"{student_name}_{doc_type}.png"
            img_path = os.path.join(student_folder, filename)
            lbl.image._PhotoImage__photo.write(img_path, format='png')

def excel_to_pdf(excel_file,output_pdf):
    
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            wb = excel.Workbooks.Open(excel_file)
            
            output_pdf_full_path = output_pdf
            print(output_pdf_full_path)
            
            wb.ActiveSheet.ExportAsFixedFormat(0, output_pdf_full_path)
            
            print("PDF successfully created at:", output_pdf_full_path)



            return output_pdf_full_path


Label(root, text="For any Query : info@mvcec.com", width=10, height=3, bg="#f0687c", anchor='e').pack(side=TOP, fill=X)

header_frame = Frame(root, bg="#f0687c")
header_frame.pack(side=TOP, fill=X)

# Load the college header image
header_image_path = "C:/Users/abin/OneDrive/Desktop/F/REimages/College name.png"
header_image = Image.open(header_image_path)
header_image = header_image.resize((580, 51))  # Resize the image as needed
header_photo = ImageTk.PhotoImage(header_image)

# Add the college header image to the frame, centered
header_label = Label(header_frame, image=header_photo, bg="#f0687c")
header_label.image = header_photo  # Keep a reference to avoid garbage collection
header_label.pack(side=LEFT, expand=True)





Label(root, text="REGISTRATION FORM", width=10, height=2, bg="#c36464", fg='#fff', font='arial 20 bold').pack(side=TOP, fill=X)

Label(root, text="Registration No:", font="arial 13", fg="white", bg=background).place(x=30, y=150)
Label(root, text="Date:", font="arial 13", fg="white", bg=background).place(x=300, y=150)

Registration = StringVar()
Date = StringVar()

reg_entry = Entry(root, textvariable=Registration, width=15, font="arial 10", state='readonly')
reg_entry.place(x=160, y=150)

def registration_no():
    file = openpyxl.load_workbook('Student_data1.xlsx')
    sheet = file.active
    row = sheet.max_row

    if row == 1:  # If there's only the header row, start with 101
        new_reg_no = "MV/CEC/101"
    else:
        last_reg_no = sheet.cell(row=row, column=1).value
        if isinstance(last_reg_no, int):
            last_reg_no = f"MV/CEC/{last_reg_no}"
        last_number = int(last_reg_no.split('/')[-1])
        new_reg_no = f"MV/CEC/{last_number + 1}"

    Registration.set(new_reg_no)

registration_no()

today = date.today()
d1 = today.strftime("%d/%m/%Y")
date_entry = Entry(root, textvariable=Date, width=15, font="arial 10", state='readonly')
date_entry.place(x=350, y=150)
Date.set(d1)

obj = LabelFrame(root, text="Student's Details", font=20, bd=2, width=1470, fg=framefg, bg=framebg, height=570, relief=GROOVE)
obj.place(x=30, y=200)

Label(obj, text="Name:", font="arial 15", bg=framebg, fg=framefg).place(x=30, y=20)
Name = StringVar()
name_entry = Entry(obj, textvariable=Name, width=20, font="arial 15")
name_entry.place(x=220, y=20)

Label(obj, text="Date of Birth:", font="arial 15", bg=framebg, fg=framefg).place(x=30, y=65)
DOB = StringVar()

dob_entry = Entry(obj, textvariable=DOB, width=20, font="arial 15")
dob_entry.place(x=220, y=65)

def get_selected_date():
    cal = Calendar(root, selectmode='day', year=2005, month=5, day=22,
                   mindate=date(1985, 1, 1), maxdate=date(2011, 1, 1),
                   width=10, height=6)
    cal.place(x=220, y=100)
    
    def update_dob():
        selected_date = cal.get_date()
        DOB.set(selected_date)
        dob_entry.delete(0, END)
        dob_entry.insert(0, selected_date)
        cal.destroy()
    
    Button(root, text="Confirm Date", command=update_dob, font="arial 13").place(x=585, y=282)

Button(obj, text="Select Date", command=get_selected_date, font="arial 13").place(x=450, y=62)

Label(obj, text="Gender:", font="arial 15", bg=framebg, fg=framefg).place(x=30, y=110)
gender = StringVar(value="Male")

R1 = Radiobutton(obj, text="Male", variable=gender, value="Male", bg=framebg, fg=framefg, font="arial 15")
R1.place(x=220, y=110)

R2 = Radiobutton(obj, text="Female", variable=gender, value="Female", bg=framebg, fg=framefg, font="arial 15")
R2.place(x=295, y=110)

gender.set("Male")

Label(obj, text="Course Title:", font="arial 15", bg=framebg, fg=framefg).place(x=30, y=155)

course_options = ["Introduction to Data Science", "Basic Data Science", "Advanced Data Science", "Java Basic", "Java Advanced", "Android", "DBMS(SQL+PLSQL)"]
course_combobox = Combobox(obj, values=course_options, font="arial 15", state='readonly')
course_combobox.place(x=220, y=155)
course_combobox.set("Select a course")

Email_id = StringVar()
Label(obj, text="Email-id: ", font="arial 15", bg=framebg, fg=framefg).place(x=30, y=200)
Email_id_entry = Entry(obj, textvariable=Email_id, width=20, font="arial 15")
Email_id_entry.place(x=220, y=200)

Label(obj, text="Students Mobile No: ", font="arial 15", bg=framebg, fg=framefg).place(x=30, y=255)
Student_Mobile = StringVar()
S_entry = Entry(obj, textvariable=Student_Mobile, width=20, font="arial 15")
S_entry.place(x=220, y=255)

Label(obj, text="Parent Contact No : ", font="arial 15", bg=framebg, fg=framefg).place(x=30, y=300)
Parent_Contact = StringVar()
PC_entry = Entry(obj, textvariable=Parent_Contact, width=20, font="arial 15")
PC_entry.place(x=220, y=300)

Label(obj, text="Qualification : ", font="arial 15", bg=framebg, fg=framefg).place(x=30, y=355)
Quali_student = StringVar()
QS_entry = Entry(obj, textvariable=Quali_student, width=20, font="arial 15")
QS_entry.place(x=220, y=355)

Label(obj, text="Address : ", font="arial 15", bg=framebg, fg=framefg).place(x=30, y=400)
Address_text = Text(obj, width=35, height=5, font=("Arial", 15))
Address_text.place(x=220, y=400)

def showimage(lbl):
    global filename
    global img
    filename = filedialog.askopenfilename(initialdir='C:/Users/abin/OneDrive/Desktop/F/images/',
                                          title="Select Image file")
                                        #   filetype=(("JPG File", "*.jpg"),
                                        #             ("PNG File", "*.png"),
                                        #             ("All files", "*.*")))
    if filename:
        img = Image.open(filename)
        photo = ImageTk.PhotoImage(img)
        lbl.config(image=photo)
        lbl.image = photo

def create_frame_with_button(x, y, image_path):
    f = Frame(root, bd=3, bg="black", width=200, height=200, relief=GROOVE)
    f.place(x=x, y=y)
    img = Image.open(image_path)
    resized_image = img.resize((190, 190))
    photo = ImageTk.PhotoImage(resized_image)
    lbl = Label(f, bg="black", image=photo)
    lbl.image = photo
    lbl.place(x=0, y=0)
    upload_button = Button(root, text="Upload", width=19, height=2, font="arial 12 bold", bg="lightblue", command=lambda: showimage(lbl))
    upload_button.place(x=x, y=y + 200)
    return lbl

frame1_lbl = create_frame_with_button(750, 200, "C:/Users/abin/OneDrive/Desktop/F/REimages/upload photo.png")
frame2_lbl = create_frame_with_button(1000, 200, "C:/Users/abin/OneDrive/Desktop/F/REimages/signature_logo_round 1.png")
frame3_lbl = create_frame_with_button(1000, 500, "C:/Users/abin/OneDrive/Desktop/F/REimages/Aadhar1.png")
frame4_lbl = create_frame_with_button(750, 500, "C:/Users/abin/OneDrive/Desktop/F/REimages/electricity bill.png")

saveButton = Button(root, text="Save", width=19, height=2, font="arial 12 bold", bg="lightgreen", command=Save)
saveButton.place(x=1270, y=370)

Button(root, text="RESET", width=19, height=2, font="arial 12 bold", bg="lightpink", command=Clear).place(x=1270, y=470)

Button(root, text="Exit", width=19, height=2, font="arial 12 bold", bg="red", command=Exit).place(x=1270, y=570)

root.mainloop()
